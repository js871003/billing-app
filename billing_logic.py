"""
과금 자동화 로직 v2.1
- 28.02 / 29.02 계약 구분 (2024 / 2025)
- 한결교육 프로모션 지원
- 디렉토리 직접판매 분리 (별도 양식)
- 새 거래명세서 양식: 2024_주N회 / 2025_주N회 / 2025_한결프로모션
"""

import pandas as pd
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime


# ========== 기준파일 로딩 ==========

def load_price_criteria(criteria_path):
    """청구요금 기준파일 로드 → 사이트아이디 → 요금 매핑"""
    xls = pd.ExcelFile(criteria_path)
    price_lookup = {}
    directory_sites = set()

    # 1) 직접판매_디렉토리
    for sheet in xls.sheet_names:
        if '직접판매' in sheet and '디렉토리' in sheet:
            df = pd.read_excel(xls, sheet_name=sheet)
            for _, row in df.iterrows():
                site_id = row['사이트아이디']
                directory_sites.add(site_id)
                price_lookup[site_id] = {
                    'price': row['요금'],
                    'price_after_promo': row['요금'],
                    'source': sheet,
                    'contract_year': '디렉토리',
                    '담당지사': _normalize_jisa(row['담당지사']),
                    '요금제': row['요금제'],
                    'promo_end': None,
                }

    # 2) 총판판매_28.02월 종료 (2024 계약분)
    for sheet in xls.sheet_names:
        if '총판판매' in sheet and '28.02' in sheet:
            df = pd.read_excel(xls, sheet_name=sheet)
            price_col = [c for c in df.columns if '요금' in str(c) and c != '요금제'][0]
            for _, row in df.iterrows():
                site_id = row['사이트아이디']
                price_lookup[site_id] = {
                    'price': row[price_col],
                    'price_after_promo': row[price_col],
                    'source': sheet,
                    'contract_year': '2024',
                    '담당지사': _normalize_jisa(row['담당지사']),
                    '요금제': row['요금제'],
                    'promo_end': None,
                }

    # 3) 총판판매_29.02월 종료 (2025 계약분, 프로모션 포함)
    for sheet in xls.sheet_names:
        if '총판판매' in sheet and '29.02' in sheet:
            df = pd.read_excel(xls, sheet_name=sheet)
            price_cols = [c for c in df.columns if '요금' in str(c) and c != '요금제']
            price_cols_sorted = sorted(price_cols)

            for _, row in df.iterrows():
                site_id = row['사이트아이디']
                if site_id in directory_sites:
                    continue
                if len(price_cols_sorted) >= 2:
                    price_promo = row[price_cols_sorted[0]]
                    price_normal = row[price_cols_sorted[1]]
                    promo_end = _extract_date_from_col(price_cols_sorted[0])
                else:
                    price_promo = row[price_cols_sorted[0]]
                    price_normal = price_promo
                    promo_end = None

                price_lookup[site_id] = {
                    'price': price_promo,
                    'price_after_promo': price_normal,
                    'source': sheet,
                    'contract_year': '2025',
                    '담당지사': _normalize_jisa(row['담당지사']),
                    '요금제': row['요금제'],
                    'promo_end': promo_end,
                }

    return price_lookup, directory_sites


def _normalize_jisa(value):
    if isinstance(value, datetime):
        return '오월오일'
    return str(value).strip()


def _extract_date_from_col(col_name):
    import re
    match = re.search(r'~(\d{4})\.(\d{2})', col_name)
    if match:
        return (int(match.group(1)), int(match.group(2)))
    return None


def is_promo_active(promo_end, billing_year, billing_month):
    if promo_end is None:
        return True
    end_year, end_month = promo_end
    if billing_year < end_year:
        return True
    if billing_year == end_year and billing_month <= end_month:
        return True
    return False


# ========== 요금제 매핑 ==========

def classify_plan(plan_str):
    """요일 표기 → 주N회"""
    if plan_str in ['주3회', '주5회']:
        return plan_str
    days = [d.strip() for d in str(plan_str).split(',')]
    return f"주{len(days)}회"


# ========== 거래명세서용 카테고리 분류 ==========

def categorize_for_invoice(row):
    """
    각 행을 거래명세서 카테고리로 분류.
    
    우선순위:
    1. contract_year가 명확히 있으면 그것을 신뢰
    2. 없으면 가격 + 담당지사로 추정
    
    가격별 분류 규칙:
    - 72,000원 → 디렉토리 (고유)
    - 54,000원 → 2024_주5회/3년 (고유)
    - 43,200원 → 2024_주3회/3년 (고유)
    - 95,000원 → 2025_주5회/3년 (고유)
    - 83,000원 → 2025_주3회/3년 (고유)
    - 60,000원 → 2024인지 2025 한결프로모션인지는 contract_year/담당지사로 구분
    - 48,000원 → 2024인지 2025 한결프로모션인지는 contract_year/담당지사로 구분
    """
    price = int(row.get('요금', 0) or 0)
    jisa = str(row.get('담당지사', '') or '').strip()
    contract_year = str(row.get('contract_year', '') or '').strip()

    # 디렉토리 (72,000원 고유)
    if price == 72000 or contract_year == '디렉토리':
        return {'category': '디렉토리', '약정': 'BASIC(3)', 'sort_order': 999}

    # 2024 고유 가격
    if price == 54000:
        return {'category': '2024_주5회', '약정': '3년', 'sort_order': 2}
    if price == 43200:
        return {'category': '2024_주3회', '약정': '3년', 'sort_order': 4}

    # 2025 고유 가격
    if price == 95000:
        return {'category': '2025_주5회', '약정': '3년', 'sort_order': 5}
    if price == 83000:
        return {'category': '2025_주3회', '약정': '3년', 'sort_order': 6}

    # 60,000원: contract_year 우선, 없으면 담당지사로 추정
    if price == 60000:
        if contract_year == '2025':
            return {'category': '2025_한결_주5회', '약정': '3년', 'sort_order': 7}
        if contract_year == '2024':
            return {'category': '2024_주5회', '약정': '1년', 'sort_order': 1}
        # contract_year 없음 → 담당지사로 추정
        if jisa == '한결교육':
            return {'category': '2025_한결_주5회', '약정': '3년', 'sort_order': 7}
        return {'category': '2024_주5회', '약정': '1년', 'sort_order': 1}

    # 48,000원: contract_year 우선, 없으면 담당지사로 추정
    if price == 48000:
        if contract_year == '2025':
            return {'category': '2025_한결_주3회', '약정': '3년', 'sort_order': 8}
        if contract_year == '2024':
            return {'category': '2024_주3회', '약정': '1년', 'sort_order': 3}
        # contract_year 없음 → 담당지사로 추정
        if jisa == '한결교육':
            return {'category': '2025_한결_주3회', '약정': '3년', 'sort_order': 8}
        return {'category': '2024_주3회', '약정': '1년', 'sort_order': 3}

    return {'category': '기타', '약정': '', 'sort_order': 999}


# ========== 1단계: 과금 판단 ==========

def process_billing(df):
    """과금 가능 여부 파일 처리 (O/X → 가능/불가능)"""
    df = df.copy()
    df.columns = df.columns.str.strip()
    df['요금제_분류'] = df['요금제'].apply(classify_plan)

    if set(df['과금 가능 여부'].unique()) <= {'O', 'X'}:
        df['과금 가능 여부'] = df['과금 가능 여부'].map({'O': '가능', 'X': '불가능'})

    stats = {
        'total_count': len(df),
        'ok_count': (df['과금 가능 여부'] == '가능').sum(),
        'fail_count': (df['과금 가능 여부'] == '불가능').sum(),
        'review_count': (df['과금 가능 여부'] == '확인필요').sum(),
    }
    review_items = df[df['과금 가능 여부'] == '확인필요'] if stats['review_count'] > 0 else pd.DataFrame()
    return df, stats, review_items


def assign_prices(df, price_lookup, billing_year=2026, billing_month=3):
    """과금 Raw에 요금/계약구분/카테고리 컬럼 추가"""
    df = df.copy()

    def get_info(row):
        if row['과금 가능 여부'] != '가능':
            return {'price': 0, 'jisa': '', 'source': '', 'contract_year': ''}
        info = price_lookup.get(row['사이트아이디'])
        if not info:
            return {'price': 0, 'jisa': '', 'source': '', 'contract_year': ''}
        if is_promo_active(info.get('promo_end'), billing_year, billing_month):
            price = info['price']
        else:
            price = info['price_after_promo']
        return {
            'price': price,
            'jisa': info.get('담당지사', ''),
            'source': info.get('source', ''),
            'contract_year': info.get('contract_year', ''),
        }

    info_series = df.apply(get_info, axis=1)
    df['요금'] = info_series.apply(lambda x: x['price'])
    df['담당지사'] = info_series.apply(lambda x: x['jisa'])
    df['계약구분'] = info_series.apply(lambda x: x['source'])
    df['contract_year'] = info_series.apply(lambda x: x['contract_year'])
    df['비고'] = ''
    return df


# ========== 정합성 체크 ==========

def check_plan_consistency(df, price_lookup):
    mismatches = []
    df_ok = df[df['과금 가능 여부'] == '가능']
    for _, row in df_ok.iterrows():
        site_id = row['사이트아이디']
        billing_plan = row.get('요금제_분류', classify_plan(row['요금제']))
        info = price_lookup.get(site_id)
        if info and billing_plan != info['요금제']:
            mismatches.append({
                'site_id': site_id,
                'institution': row['기관명'],
                'billing_plan': billing_plan,
                'criteria_plan': info['요금제'],
            })
    return mismatches


def check_issue_list_excluded(df, criteria_path):
    xls = pd.ExcelFile(criteria_path)
    issue_sheets = [s for s in xls.sheet_names if '이슈' in s]
    if not issue_sheets:
        return 0, 0, []
    df_issue = pd.read_excel(xls, sheet_name=issue_sheets[0])
    issue_sites = set(df_issue['사이트아이디'])
    billing_sites = set(df['사이트아이디'])
    found = issue_sites & billing_sites
    return len(issue_sites) - len(found), len(issue_sites), list(found)


# ========== 공급자 정보 ==========

SUPPLIER_INFO = {
    'biz_no': '870-88-02332',
    'company': '플레이태그',
    'ceo': '박현수',
    'phone': '02-553-0214',
    'address': '서울 강남구 삼성동 천해빌딩 10층',
    'bank': '하나은행 / 플레이태그주식회사 / 403-910059-30704',
}


def split_by_directory(df, directory_sites):
    """과금 Raw → 총판용 / 디렉토리용 분리"""
    mask = df['사이트아이디'].isin(directory_sites)
    df_directory = df[mask].copy()
    df_wholesale = df[~mask].copy()
    return df_wholesale, df_directory


# ========== 거래명세서 양식 헬퍼 ==========

def _set_range_border(ws, cell_range, border):
    """병합 셀 범위의 외곽선만 설정 (내부 분할선 없음).
    openpyxl에서 병합 셀의 내부 셀에도 border를 설정하면
    LibreOffice/Excel에서 내부 분할선이 보이는 문제를 해결합니다.
    """
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)

    # border 객체에서 side 추출 (모든 방향이 같다고 가정)
    side = border.left

    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            cell.border = Border(
                left=side if col == min_col else None,
                right=side if col == max_col else None,
                top=side if row == min_row else None,
                bottom=side if row == max_row else None,
            )


def _outer_border_range(ws, cell_range, side):
    """범위의 외곽선만 medium 두께로 그리기"""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            existing = cell.border
            new_border = Border(
                left=side if col == min_col else existing.left,
                right=side if col == max_col else existing.right,
                top=side if row == min_row else existing.top,
                bottom=side if row == max_row else existing.bottom,
            )
            cell.border = new_border


def _fill_range(ws, cell_range, fill):
    """범위 내 모든 셀에 fill 적용"""
    from openpyxl.utils import range_boundaries
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            ws.cell(row=row, column=col).fill = fill


def _setup_invoice_common(ws, recipient_name, recipient_info, billing_date,
                           supply=0, vat=0, total=0):
    """거래명세서 공통 영역: 헤더(Row 1~9), 합계(Row 33~37)
    데이터 테이블(Row 10~32)은 호출 측에서 설정
    supply/vat/total: 미리 계산된 값 (수식 대신 직접 표시)"""

    # 그리드 라인 숨기기 (빈 셀 사이 구분선 제거)
    ws.sheet_view.showGridLines = False

    # 인쇄 설정
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    ws.page_setup.orientation = 'portrait'
    ws.page_setup.fitToPage = True
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 1
    ws.print_area = 'A1:M37'
    ws.page_margins.left = 0.4
    ws.page_margins.right = 0.4
    ws.page_margins.top = 0.4
    ws.page_margins.bottom = 0.4

    # 스타일
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    font_title = Font(name='Dotum', size=22, bold=True)
    font_normal = Font(name='Dotum', size=10)
    font_normal_bold = Font(name='Dotum', size=10, bold=True)
    font_recipient = Font(name='Dotum', size=12, bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')

    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)

    # === Row 1: 빈 행 ===
    ws.row_dimensions[1].height = 12

    # === Row 2: 제목 ===
    ws.row_dimensions[2].height = 38
    ws.merge_cells('A2:M2')
    ws['A2'] = '거 래 명 세 서'
    ws['A2'].font = font_title
    ws['A2'].alignment = Alignment(horizontal='center', vertical='center')

    # === Row 3: 빈 행 ===
    ws.row_dimensions[3].height = 12

    # === Row 4: 날짜 ===
    ws.row_dimensions[4].height = 22
    ws['A4'] = '날짜 :'
    ws['A4'].font = font_normal_bold
    ws['A4'].alignment = left_align
    ws.merge_cells('B4:C4')
    ws['B4'] = billing_date if billing_date else datetime.now()
    ws['B4'].font = font_normal_bold
    ws['B4'].number_format = 'YYYY"년" M"월" D"일"'
    ws['B4'].alignment = left_align

    # === Row 5~8: 공급자/수신자 정보 ===
    for r in range(5, 9):
        ws.row_dimensions[r].height = 24

    # 공/급/자 세로 병합 셀
    ws.merge_cells('D5:D8')
    ws['D5'] = '공\n급\n자'
    ws['D5'].font = font_normal_bold
    ws['D5'].alignment = center
    _set_range_border(ws, 'D5:D8', thin)
    _fill_range(ws, 'D5:D8', header_fill)

    # R5: 사업자등록번호
    ws.merge_cells('E5:G5')
    ws['E5'] = '사업자등록번호'
    ws['E5'].font = font_normal_bold
    ws['E5'].alignment = center
    _set_range_border(ws, 'E5:G5', thin)
    _fill_range(ws, 'E5:G5', header_fill)

    ws.merge_cells('H5:M5')
    ws['H5'] = SUPPLIER_INFO['biz_no']
    ws['H5'].font = font_normal
    ws['H5'].alignment = center
    _set_range_border(ws, 'H5:M5', thin)

    # R6: 수신처(좌) + 상호 + 대표자(우)
    ws.merge_cells('A6:C6')
    ws['A6'] = f' {recipient_name} 귀하'
    ws['A6'].font = font_recipient
    ws['A6'].alignment = Alignment(horizontal='left', vertical='center')
    # 수신처 이름 아래 밑줄
    ws['A6'].border = Border(bottom=Side(style='thin'))
    ws['B6'].border = Border(bottom=Side(style='thin'))
    ws['C6'].border = Border(bottom=Side(style='thin'))

    ws.merge_cells('E6:G6')
    ws['E6'] = '상호'
    ws['E6'].font = font_normal_bold
    ws['E6'].alignment = center
    _set_range_border(ws, 'E6:G6', thin)
    _fill_range(ws, 'E6:G6', header_fill)

    ws.merge_cells('H6:J6')
    ws['H6'] = SUPPLIER_INFO['company']
    ws['H6'].font = font_normal
    ws['H6'].alignment = center
    _set_range_border(ws, 'H6:J6', thin)

    ws.merge_cells('K6:L6')
    ws['K6'] = '대표자'
    ws['K6'].font = font_normal_bold
    ws['K6'].alignment = center
    _set_range_border(ws, 'K6:L6', thin)
    _fill_range(ws, 'K6:L6', header_fill)

    ws['M6'] = f' {SUPPLIER_INFO["ceo"]} (인)'
    ws['M6'].font = font_normal
    ws['M6'].alignment = left_align
    ws['M6'].border = thin

    # 도장 이미지 삽입 (대표자 이름 옆)
    import os
    stamp_path = os.path.join(
        os.path.dirname(os.path.abspath(__file__)), 'stamp.png'
    )
    if os.path.exists(stamp_path):
        try:
            from openpyxl.drawing.image import Image as XLImage
            from openpyxl.drawing.spreadsheet_drawing import OneCellAnchor
            from openpyxl.drawing.xdr import XDRPositiveSize2D
            from openpyxl.utils.units import pixels_to_EMU

            stamp_img = XLImage(stamp_path)
            stamp_img.width = 45
            stamp_img.height = 45
            # M6 우측에 위치시키기
            ws.add_image(stamp_img, 'M6')
        except Exception:
            pass  # 도장 로드 실패 시 무시

    # R7~8: 수신처 정보(좌) + 전화번호/주소(우)
    addr_text = recipient_info.get('address', '')
    biz_no = recipient_info.get('biz_no', '')
    email = recipient_info.get('email', '')
    info_lines = []
    if addr_text:
        info_lines.append(f'주소 : {addr_text}')
    if biz_no:
        info_lines.append(f'사업자 번호 : {biz_no}')
    if email:
        info_lines.append(f'이메일 : {email}')
    left_info = '\n'.join(info_lines)

    ws.merge_cells('A7:C8')
    ws['A7'] = left_info
    ws['A7'].font = font_normal
    ws['A7'].alignment = Alignment(vertical='center', wrap_text=True, horizontal='left')

    # R7: 전화번호
    ws.merge_cells('E7:G7')
    ws['E7'] = '전화번호'
    ws['E7'].font = font_normal_bold
    ws['E7'].alignment = center
    _set_range_border(ws, 'E7:G7', thin)
    _fill_range(ws, 'E7:G7', header_fill)

    ws.merge_cells('H7:M7')
    ws['H7'] = SUPPLIER_INFO['phone']
    ws['H7'].font = font_normal
    ws['H7'].alignment = center
    _set_range_border(ws, 'H7:M7', thin)

    # R8: 주소
    ws.merge_cells('E8:G8')
    ws['E8'] = '주소'
    ws['E8'].font = font_normal_bold
    ws['E8'].alignment = center
    _set_range_border(ws, 'E8:G8', thin)
    _fill_range(ws, 'E8:G8', header_fill)

    ws.merge_cells('H8:M8')
    ws['H8'] = SUPPLIER_INFO['address']
    ws['H8'].font = font_normal
    ws['H8'].alignment = center
    _set_range_border(ws, 'H8:M8', thin)

    # === Row 9: 합계금액 ===
    ws.row_dimensions[9].height = 30
    ws.merge_cells('A9:B9')
    ws['A9'] = '합계금액 :'
    ws['A9'].font = font_recipient
    ws['A9'].alignment = Alignment(horizontal='right', vertical='center')

    ws.merge_cells('C9:H9')
    ws['C9'] = total
    ws['C9'].font = font_recipient
    ws['C9'].number_format = '#,##0" 원정"'
    ws['C9'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('I9:J9')
    ws['I9'] = '₩'
    ws['I9'].font = font_recipient
    ws['I9'].alignment = Alignment(horizontal='center', vertical='center')

    ws.merge_cells('K9:M9')
    ws['K9'] = total
    ws['K9'].font = font_recipient
    ws['K9'].number_format = '#,##0'
    ws['K9'].alignment = Alignment(horizontal='right', vertical='center')

    # === Row 33: 작은 간격 ===
    ws.row_dimensions[33].height = 4

    # === Row 34~36: 비고 + 합계 영역 ===
    sum_fill = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')
    for r in [34, 35, 36]:
        ws.row_dimensions[r].height = 24

    # 비고 (3행 병합)
    ws.merge_cells('A34:G36')
    ws['A34'] = '비 고 : \n1) 상세 데이터 별도 제공'
    ws['A34'].font = font_normal_bold
    ws['A34'].alignment = Alignment(vertical='top', wrap_text=True, horizontal='left')
    _set_range_border(ws, 'A34:G36', thin)

    # 공급가
    ws.merge_cells('H34:J34')
    ws['H34'] = '공 급 가'
    ws['H34'].font = font_normal_bold
    ws['H34'].alignment = center
    _set_range_border(ws, 'H34:J34', thin)
    _fill_range(ws, 'H34:J34', sum_fill)

    ws.merge_cells('K34:M34')
    ws['K34'] = supply
    ws['K34'].font = font_normal_bold
    ws['K34'].alignment = Alignment(horizontal='right', vertical='center')
    ws['K34'].number_format = '#,##0'
    _set_range_border(ws, 'K34:M34', thin)

    # 부가세
    ws.merge_cells('H35:J35')
    ws['H35'] = '부가세'
    ws['H35'].font = font_normal_bold
    ws['H35'].alignment = center
    _set_range_border(ws, 'H35:J35', thin)
    _fill_range(ws, 'H35:J35', sum_fill)

    ws.merge_cells('K35:M35')
    ws['K35'] = vat
    ws['K35'].font = font_normal_bold
    ws['K35'].alignment = Alignment(horizontal='right', vertical='center')
    ws['K35'].number_format = '#,##0'
    _set_range_border(ws, 'K35:M35', thin)

    # 합계금액
    ws.merge_cells('H36:J36')
    ws['H36'] = '합 계 금 액'
    ws['H36'].font = font_normal_bold
    ws['H36'].alignment = center
    _set_range_border(ws, 'H36:J36', thin)
    _fill_range(ws, 'H36:J36', sum_fill)

    ws.merge_cells('K36:M36')
    ws['K36'] = total
    ws['K36'].font = Font(name='Dotum', size=11, bold=True)
    ws['K36'].alignment = Alignment(horizontal='right', vertical='center')
    ws['K36'].number_format = '#,##0'
    _set_range_border(ws, 'K36:M36', thin)

    # === Row 37: 입금정보 ===
    ws.row_dimensions[37].height = 30
    ws.merge_cells('A37:M37')
    ws['A37'] = f'   입금정보 : {SUPPLIER_INFO["bank"]}'
    ws['A37'].font = font_normal_bold
    ws['A37'].alignment = Alignment(vertical='center', horizontal='left')
    _set_range_border(ws, 'A37:M37', thin)


def _draw_data_table_header(ws, col_c_header='요금제', col_d_header='약정'):
    """Row 10 데이터 테이블 헤더"""
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    font_header = Font(name='Dotum', size=10, bold=True)
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)

    ws.row_dimensions[10].height = 26

    # A: 이용월
    ws['A10'] = '이용월'
    ws['A10'].font = font_header
    ws['A10'].alignment = center
    ws['A10'].fill = header_fill
    ws['A10'].border = thin

    # B: 대리점
    ws['B10'] = '대리점'
    ws['B10'].font = font_header
    ws['B10'].alignment = center
    ws['B10'].fill = header_fill
    ws['B10'].border = thin

    # C: 요금제 (또는 기관명)
    ws['C10'] = col_c_header
    ws['C10'].font = font_header
    ws['C10'].alignment = center
    ws['C10'].fill = header_fill
    ws['C10'].border = thin

    # D-F: 약정 (또는 구분)
    ws.merge_cells('D10:F10')
    ws['D10'] = col_d_header
    ws['D10'].font = font_header
    ws['D10'].alignment = center
    _set_range_border(ws, 'D10:F10', thin)
    _fill_range(ws, 'D10:F10', header_fill)

    # G: 수량
    ws['G10'] = '수량'
    ws['G10'].font = font_header
    ws['G10'].alignment = center
    ws['G10'].fill = header_fill
    ws['G10'].border = thin

    # H: 요율
    ws['H10'] = '요율'
    ws['H10'].font = font_header
    ws['H10'].alignment = center
    ws['H10'].fill = header_fill
    ws['H10'].border = thin

    # I-J: 단가
    ws.merge_cells('I10:J10')
    ws['I10'] = '단가'
    ws['I10'].font = font_header
    ws['I10'].alignment = center
    _set_range_border(ws, 'I10:J10', thin)
    _fill_range(ws, 'I10:J10', header_fill)

    # K-M: 금 액
    ws.merge_cells('K10:M10')
    ws['K10'] = '금 액'
    ws['K10'].font = font_header
    ws['K10'].alignment = center
    _set_range_border(ws, 'K10:M10', thin)
    _fill_range(ws, 'K10:M10', header_fill)


def _draw_data_row(ws, row_idx, billing_month, agency, col_c, col_d,
                   qty, rate_text, unit_price):
    """데이터 행 한 줄 그리기 (모든 병합 셀 border 처리)"""
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    font_data = Font(name='Dotum', size=10, bold=True)
    font_data_normal = Font(name='Dotum', size=9)
    font_num = Font(name='Malgun Gothic', size=10)
    center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    right = Alignment(horizontal='right', vertical='center')

    ws.row_dimensions[row_idx].height = 22

    # A: 이용월
    ws[f'A{row_idx}'] = billing_month
    ws[f'A{row_idx}'].font = font_data
    ws[f'A{row_idx}'].alignment = center
    ws[f'A{row_idx}'].border = thin

    # B: 대리점
    ws[f'B{row_idx}'] = agency
    ws[f'B{row_idx}'].font = font_data_normal
    ws[f'B{row_idx}'].alignment = center
    ws[f'B{row_idx}'].border = thin

    # C: 카테고리/기관명
    ws[f'C{row_idx}'] = col_c
    ws[f'C{row_idx}'].font = font_data
    ws[f'C{row_idx}'].alignment = center
    ws[f'C{row_idx}'].border = thin

    # D-F: 약정/구분
    ws.merge_cells(f'D{row_idx}:F{row_idx}')
    ws[f'D{row_idx}'] = col_d
    ws[f'D{row_idx}'].font = font_data
    ws[f'D{row_idx}'].alignment = center
    _set_range_border(ws, f'D{row_idx}:F{row_idx}', thin)

    # G: 수량
    ws[f'G{row_idx}'] = qty
    ws[f'G{row_idx}'].font = font_data
    ws[f'G{row_idx}'].alignment = center
    ws[f'G{row_idx}'].border = thin

    # H: 요율 (텍스트로)
    ws[f'H{row_idx}'] = rate_text
    ws[f'H{row_idx}'].font = font_data
    ws[f'H{row_idx}'].alignment = center
    ws[f'H{row_idx}'].border = thin

    # I-J: 단가
    ws.merge_cells(f'I{row_idx}:J{row_idx}')
    ws[f'I{row_idx}'] = unit_price
    ws[f'I{row_idx}'].font = font_num
    ws[f'I{row_idx}'].alignment = center
    ws[f'I{row_idx}'].number_format = '#,##0'
    _set_range_border(ws, f'I{row_idx}:J{row_idx}', thin)

    # K-M: 금액 (계산식)
    rate_decimal = 1.0 if rate_text == '100%' else 0.5
    ws.merge_cells(f'K{row_idx}:M{row_idx}')
    ws[f'K{row_idx}'] = int(unit_price * qty * rate_decimal)
    ws[f'K{row_idx}'].font = font_data
    ws[f'K{row_idx}'].alignment = right
    ws[f'K{row_idx}'].number_format = '#,##0'
    _set_range_border(ws, f'K{row_idx}:M{row_idx}', thin)


def _draw_empty_row(ws, row_idx):
    """빈 데이터 행 (테두리만)"""
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    ws.row_dimensions[row_idx].height = 22
    for col in ['A', 'B', 'C', 'G', 'H']:
        ws[f'{col}{row_idx}'].border = thin
    ws.merge_cells(f'D{row_idx}:F{row_idx}')
    _set_range_border(ws, f'D{row_idx}:F{row_idx}', thin)
    ws.merge_cells(f'I{row_idx}:J{row_idx}')
    _set_range_border(ws, f'I{row_idx}:J{row_idx}', thin)
    ws.merge_cells(f'K{row_idx}:M{row_idx}')
    _set_range_border(ws, f'K{row_idx}:M{row_idx}', thin)


# ========== 총판 거래명세서 ==========

def create_invoice_excel(df_raw, recipient_name, billing_month, recipient_info=None,
                          billing_date=None):
    """
    총판 거래명세서 생성 (reference 1월 PDF 양식 기준)
    카테고리: 2024_주5회 / 2024_주3회 / 2025_주5회 / 2025_주3회 / 2025_한결_주5회 / 2025_한결_주3회
    """
    if recipient_info is None:
        recipient_info = {'address': '', 'biz_no': '', 'email': ''}

    billing_ok = df_raw[df_raw['과금 가능 여부'] == '가능'].copy()

    cats = billing_ok.apply(categorize_for_invoice, axis=1)
    billing_ok['category'] = cats.apply(lambda x: x['category'])
    billing_ok['약정_label'] = cats.apply(lambda x: x['약정'])
    billing_ok['sort_order'] = cats.apply(lambda x: x['sort_order'])

    grouped = billing_ok.groupby(
        ['sort_order', 'category', '약정_label', '요금']
    ).size().reset_index(name='수량')
    grouped = grouped.sort_values(['sort_order', '요금'], ascending=[True, False])

    # 미리 합계 계산 (수식 대신 값 사용)
    supply = int(sum(int(g['요금']) * int(g['수량']) for _, g in grouped.iterrows()))
    vat = int(supply * 0.1)
    total = supply + vat

    wb = Workbook()
    ws = wb.active
    ws.title = '거래명세서'

    # 열 너비 (총판: C 컬럼 더 좁게, 요금제 카테고리 라벨 길이에 맞춤)
    col_widths = {'A': 8.86, 'B': 17.57, 'C': 32.0, 'D': 5.43, 'E': 1.86,
                  'F': 13.0, 'G': 10.71, 'H': 8.86, 'I': 3.14, 'J': 9.0,
                  'K': 2.14, 'L': 5.14, 'M': 16.43}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 공통 영역 (미리 계산한 합계 전달)
    _setup_invoice_common(ws, recipient_name, recipient_info, billing_date,
                          supply=supply, vat=vat, total=total)

    # 데이터 테이블 헤더
    _draw_data_table_header(ws, col_c_header='요금제', col_d_header='약정')

    # 데이터 행
    data_start = 11
    data_end = 32
    row_idx = data_start

    for _, grp in grouped.iterrows():
        if row_idx > data_end:
            break
        _draw_data_row(
            ws, row_idx,
            billing_month=billing_month,
            agency=recipient_name,
            col_c=grp['category'],
            col_d=grp['약정_label'],
            qty=int(grp['수량']),
            rate_text='100%',
            unit_price=int(grp['요금'])
        )
        row_idx += 1

    # 빈 행
    for r in range(row_idx, data_end + 1):
        _draw_empty_row(ws, r)

    # 외곽선
    thick = Side(style='medium')
    _outer_border_range(ws, 'A2:M37', thick)

    return wb


# ========== 디렉토리 거래명세서 ==========

def create_directory_invoice_excel(df_directory, recipient_name='(주)디렉토리',
                                    billing_month='', recipient_info=None,
                                    billing_date=None):
    """
    디렉토리 거래명세서 생성 (기관별 그룹핑, 50% 요율)
    각 행: 기관명 / BASIC(3) / 수량 / 50% / 72,000원 / 금액
    """
    if recipient_info is None:
        recipient_info = {'address': '', 'biz_no': '', 'email': ''}

    billing_ok = df_directory[df_directory['과금 가능 여부'] == '가능'].copy()

    grouped = billing_ok.groupby('기관명').agg(
        수량=('사이트아이디', 'count'),
        요금=('요금', 'first'),
    ).reset_index()
    grouped = grouped.sort_values('수량', ascending=False)

    # 미리 합계 계산 (50% 요율)
    supply = int(sum(int(g['요금']) * int(g['수량']) * 0.5 for _, g in grouped.iterrows()))
    vat = int(supply * 0.1)
    total = supply + vat

    wb = Workbook()
    ws = wb.active
    ws.title = '거래명세서'

    # 열 너비 (디렉토리: C 컬럼 = 기관명)
    col_widths = {'A': 8.86, 'B': 13.0, 'C': 25.0, 'D': 5.43, 'E': 1.86,
                  'F': 13.0, 'G': 10.71, 'H': 8.86, 'I': 3.14, 'J': 9.0,
                  'K': 2.14, 'L': 5.14, 'M': 16.43}
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width

    # 공통 영역
    _setup_invoice_common(ws, recipient_name, recipient_info, billing_date,
                          supply=supply, vat=vat, total=total)

    # 데이터 테이블 헤더 (기관명 / 구분)
    _draw_data_table_header(ws, col_c_header='기관명', col_d_header='구분')

    # 데이터 행
    data_start = 11
    data_end = 32
    row_idx = data_start

    for _, grp in grouped.iterrows():
        if row_idx > data_end:
            break
        _draw_data_row(
            ws, row_idx,
            billing_month=billing_month,
            agency='디렉토리',
            col_c=grp['기관명'],
            col_d='BASIC(3)',
            qty=int(grp['수량']),
            rate_text='50%',
            unit_price=int(grp['요금'])
        )
        row_idx += 1

    # 빈 행
    for r in range(row_idx, data_end + 1):
        _draw_empty_row(ws, r)

    # 외곽선
    thick = Side(style='medium')
    _outer_border_range(ws, 'A2:M37', thick)

    return wb


# ========== Excel → PDF 변환 ==========

def excel_to_pdf_bytes(wb):
    """
    openpyxl Workbook → PDF bytes 변환
    LibreOffice(soffice)를 사용하여 Excel 파일을 PDF로 변환합니다.
    Streamlit Cloud에서 사용하려면 packages.txt에 libreoffice 추가 필요.
    """
    import subprocess
    import tempfile
    import os

    with tempfile.TemporaryDirectory() as tmpdir:
        xlsx_path = os.path.join(tmpdir, 'invoice.xlsx')
        pdf_path = os.path.join(tmpdir, 'invoice.pdf')
        wb.save(xlsx_path)

        # LibreOffice가 설치된 경로 찾기
        soffice_cmd = None
        for cmd in ['soffice', 'libreoffice', '/usr/bin/soffice', '/usr/bin/libreoffice']:
            try:
                result = subprocess.run(
                    [cmd, '--version'],
                    capture_output=True, timeout=5
                )
                if result.returncode == 0:
                    soffice_cmd = cmd
                    break
            except (FileNotFoundError, subprocess.TimeoutExpired):
                continue

        if not soffice_cmd:
            return None  # LibreOffice 없음

        try:
            subprocess.run(
                [soffice_cmd, '--headless', '--convert-to', 'pdf',
                 '--outdir', tmpdir, xlsx_path],
                capture_output=True, timeout=60, check=True
            )
            if os.path.exists(pdf_path):
                with open(pdf_path, 'rb') as f:
                    return f.read()
        except (subprocess.TimeoutExpired, subprocess.CalledProcessError):
            return None

    return None


# ========== 별도 제공자료 ==========

def create_summary_sheet(df_raw):
    """담당지사별 요약 테이블"""
    billing_ok = df_raw[df_raw['과금 가능 여부'] == '가능'].copy()

    pivot = billing_ok.pivot_table(
        values='요금', index='담당지사', columns='요금제_분류',
        aggfunc='sum', fill_value=0
    )
    for col in ['주3회', '주5회']:
        if col not in pivot.columns:
            pivot[col] = 0

    pivot['합계'] = pivot.sum(axis=1)
    pivot = pivot[['주3회', '주5회', '합계']]
    pivot.loc['합계'] = pivot.sum()

    vat_row = pivot.loc['합계'] * 1.1
    vat_row.name = 'VAT 포함'
    pivot = pd.concat([pivot, pd.DataFrame([vat_row])])
    return pivot


def create_detail_excel(df_raw, billing_month):
    """별도 제공자료 엑셀 (요약 + Raw)"""
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = '요약'

    billing_ok = df_raw[df_raw['과금 가능 여부'] == '가능'].copy()

    pivot = billing_ok.pivot_table(
        values='요금', index='담당지사', columns='요금제_분류',
        aggfunc='sum', fill_value=0
    )
    for col in ['주3회', '주5회']:
        if col not in pivot.columns:
            pivot[col] = 0

    pivot['Grand Total'] = pivot.sum(axis=1)
    pivot = pivot[['주3회', '주5회', 'Grand Total']]
    pivot = pivot.sort_index()

    header_font = Font(name='Dotum', size=11, bold=True)
    data_font = Font(name='Dotum', size=10)
    num_fmt = '#,##0'
    thin = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    center = Alignment(horizontal='center', vertical='center')
    right = Alignment(horizontal='right', vertical='center')

    headers = ['담당지사', '주3회', '주5회', 'Grand Total']
    for col_idx, h in enumerate(headers, 1):
        cell = ws_summary.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.border = thin
        cell.alignment = center

    for row_idx, (jisa, row_data) in enumerate(pivot.iterrows(), 2):
        cell = ws_summary.cell(row=row_idx, column=1, value=jisa)
        cell.font = data_font
        cell.border = thin
        for col_idx, col_name in enumerate(['주3회', '주5회', 'Grand Total'], 2):
            val = row_data.get(col_name, 0)
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val if val != 0 else None)
            cell.font = data_font
            cell.border = thin
            cell.alignment = right
            if val != 0:
                cell.number_format = num_fmt

    ws_summary.column_dimensions['A'].width = 20
    ws_summary.column_dimensions['B'].width = 15
    ws_summary.column_dimensions['C'].width = 15
    ws_summary.column_dimensions['D'].width = 15

    month_code = billing_month.replace('.', '')
    ws_raw = wb.create_sheet(title=f'Raw_{month_code}')

    raw_columns = [
        '사이트아이디', '기관명', '반명', '가능한 일자 수', '성공 일자 수',
        '스토리라인 성공률', '담당지사', '요금제', '요금제_분류',
        '과금 가능 여부', '요금', '계약구분', 'contract_year', '비고'
    ]
    available_cols = [c for c in raw_columns if c in billing_ok.columns]

    for col_idx, h in enumerate(available_cols, 1):
        cell = ws_raw.cell(row=1, column=col_idx, value=h)
        cell.font = header_font
        cell.border = thin
        cell.alignment = center

    for row_idx, (_, row) in enumerate(billing_ok.iterrows(), 2):
        for col_idx, col_name in enumerate(available_cols, 1):
            val = row.get(col_name, '')
            if pd.isna(val):
                val = ''
            cell = ws_raw.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = thin
            if col_name == '스토리라인 성공률' and isinstance(val, (int, float)):
                cell.number_format = '0.0000'
            elif col_name == '요금' and isinstance(val, (int, float)):
                cell.number_format = num_fmt

    col_widths = [18, 15, 12, 12, 12, 15, 15, 12, 10, 12, 12, 18, 10, 10]
    for i, w in enumerate(col_widths[:len(available_cols)]):
        ws_raw.column_dimensions[get_column_letter(i + 1)].width = w

    return wb
